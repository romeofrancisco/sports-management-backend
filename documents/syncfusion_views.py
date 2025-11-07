"""
Syncfusion Document Editor API endpoints for document operations
Handles loading documents from Cloudinary and saving back to Cloudinary
"""

from django.http import JsonResponse, HttpResponseBadRequest, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
import json
import base64
import requests
from io import BytesIO
from .models import Document
import cloudinary.uploader
from cloudinary.uploader import upload as cloudinary_upload
from cloudinary.uploader import destroy as cloudinary_destroy
import io
import os
import tempfile
from django.core.files.base import ContentFile
from .utils import get_cloudinary_public_id_regex
from openpyxl import load_workbook
from cloudinary.utils import cloudinary_url
import cloudinary.api


@csrf_exempt
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_document(request):
    """
    Load a document from Cloudinary for editing in Syncfusion
    Returns the raw file for client-side processing
    Expects: { "documentId": <id> }
    Returns: Binary file download
    """
    try:
        data = json.loads(request.body)
        document_id = data.get("documentId")

        if not document_id:
            return JsonResponse({"error": "Document ID is required"}, status=400)

        # Get the document from database
        try:
            document = Document.objects.get(id=document_id)
        except Document.DoesNotExist:
            return JsonResponse({"error": "Document not found"}, status=404)

        # Check user permissions
        if not document.can_open_in_editor(request.user):
            return JsonResponse(
                {"error": "You do not have permission to view this document"}, status=403
            )

        can_edit = document.can_edit(request.user)
        is_public = document.is_in_public_folder()

        # ✅ Build Cloudinary URL manually with version to force fresh load
        public_id = document.file.name.replace("\\", "/")

        # (Optional) ensure consistent pathing if using "media/documents/"
        if not public_id.startswith("media/"):
            public_id = f"media/{public_id}"

        file_url, _ = cloudinary_url(
            public_id,
            resource_type="raw",
            version=document.version,  # ✅ Force specific version (bypasses CDN cache)
        )

        # ✅ Download the *exact* version from Cloudinary
        response = requests.get(file_url, timeout=30)
        if response.status_code != 200:
            return JsonResponse(
                {"error": "Failed to download document from storage"}, status=500
            )

        file_content = response.content
        file_extension = os.path.splitext(document.title)[1].lower()

        # Prepare form data for Syncfusion import
        files = {
            "files": (
                document.title,
                BytesIO(file_content),
                "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            )
        }

        # Try multiple Syncfusion endpoints
        conversion_urls = [
            "https://services.syncfusion.com/react/production/api/documenteditor/Import",
            "https://ej2services.syncfusion.com/production/web-services/api/documenteditor/Import",
        ]

        sfdt_data = None
        last_error = None

        for conversion_url in conversion_urls:
            try:
                conversion_response = requests.post(conversion_url, files=files, timeout=30)
                if conversion_response.status_code == 200:
                    sfdt_data = conversion_response.text
                    break
                else:
                    last_error = f"Status {conversion_response.status_code}: {conversion_response.text}"
            except Exception as e:
                last_error = str(e)
                continue

        if not sfdt_data:
            return JsonResponse(
                {
                    "error": f"Failed to convert document to SFDT format. Last error: {last_error}"
                },
                status=500,
            )

        # ✅ Return SFDT data + permission flags
        return JsonResponse(
            {
                "sfdt": sfdt_data,
                "fileName": document.title,
                "documentId": document.id,
                "canEdit": can_edit,
                "isPublic": is_public,
            }
        )

    except Exception as e:
        return JsonResponse({"error": f"Error loading document: {str(e)}"}, status=500)

@csrf_exempt
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def export_document(request):
    """
    Save document back to Cloudinary
    Expects: { "documentId": <id>, "content": <base64_encoded_data>, "fileName": <name> }
    """
    try:
        data = json.loads(request.body)
        document_id = data.get("documentId")
        content = data.get("content")
        file_name = data.get("fileName")

        if not all([document_id, content]):
            return JsonResponse(
                {"error": "Document ID and content are required"}, status=400
            )

        # Get the document from database
        document = Document.objects.filter(id=document_id).first()
        if not document:
            return JsonResponse({"error": "Document not found"}, status=404)

        # Check edit permission
        if not document.can_edit(request.user):
            if document.is_in_public_folder():
                return JsonResponse(
                    {
                        "error": "Cannot save changes to public documents",
                        "message": "This is a public document. Please make a copy to save your changes.",
                        "isPublic": True,
                    },
                    status=403,
                )
            return JsonResponse(
                {
                    "error": "You do not have permission to edit this document",
                    "message": "Only the document owner or admin can save changes to this file.",
                },
                status=403,
            )

        # Decode base64 -> bytes
        file_content = base64.b64decode(content)
        file_extension = document.file_extension or ".docx"

        # Write to a temporary file with the correct extension
        with tempfile.NamedTemporaryFile(suffix=file_extension, delete=False) as tmp_file:
            tmp_file.write(file_content)
            tmp_file_path = tmp_file.name

        # Upload updated file to Cloudinary (overwrite old one)
        public_id = document.file.name

        upload_result = cloudinary.uploader.upload(
            tmp_file_path,
            public_id=public_id,
            resource_type="raw",
            type="upload",
            overwrite=True,
            invalidate=True,
        )

        # Clean up temporary file
        os.remove(tmp_file_path)
        
        print(f"Upload result: {upload_result}")

        new_version = upload_result.get("version")
        file_url, _ = cloudinary_url(
            upload_result["public_id"],
            resource_type="raw",
            version=new_version,
        )

        # Update the document model
        document.file.name = upload_result["public_id"]
        document.version = str(new_version)
        document.save(update_fields=["file", "version"])

        return JsonResponse(
            {
                "success": True,
                "message": "Document saved successfully",
                "documentId": document.id,
                "fileUrl": file_url,
                "version": new_version,
            }
        )

    except Exception as e:
        return JsonResponse({"error": f"Error saving document: {str(e)}"}, status=500)



@csrf_exempt
@require_http_methods(["POST"])
def systemClipboard(request):
    """
    Handle system clipboard operations for Syncfusion
    This is required by Syncfusion Document Editor
    """
    try:
        data = json.loads(request.body)
        content = data.get("content", "")

        return JsonResponse({"content": content})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def restrictediting(request):
    """
    Handle restricted editing requests from Syncfusion
    """
    try:
        # For now, return empty response
        # You can implement user-specific editing restrictions here
        return JsonResponse({"canEdit": True})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


SYNCFUSION_OPEN_URL = "https://document.syncfusion.com/web-services/spreadsheet-editor/api/spreadsheet/open"


@csrf_exempt
@require_http_methods(["POST"])
def spreadsheet_open(request):
    file = request.FILES.get("file")
    if not file:
        return HttpResponseBadRequest("Missing file")

    tmp_path = None

    try:
        # Write uploaded file to temp location
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            for chunk in file.chunks():
                tmp.write(chunk)
            tmp.flush()
            tmp_path = tmp.name

        print(f"Temp file created: {tmp_path}")

        # Open file and send to Syncfusion with shorter timeout
        with open(tmp_path, "rb") as file_handle:
            files = {"file": file_handle}
            print("Sending request to Syncfusion...")
            r = requests.post(
                SYNCFUSION_OPEN_URL, files=files, timeout=15
            )  # Reduced timeout
            print(f"Syncfusion response status: {r.status_code}")

        if r.status_code != 200:
            print(f"Syncfusion error: {r.text}")
            return JsonResponse({"error": "Syncfusion conversion failed"}, status=500)

        # r.json() already returns workbook JSON
        json_data = r.json()
        print(f"Successfully converted spreadsheet")
        print(f"Response data keys: {list(json_data.keys())}")
        return JsonResponse(json_data)

    except requests.Timeout:
        print("Request to Syncfusion timed out")
        return JsonResponse(
            {"error": "Request timeout - file conversion took too long"}, status=504
        )
    except requests.RequestException as e:
        print(f"Request error: {str(e)}")
        return JsonResponse({"error": f"Network error: {str(e)}"}, status=500)
    except Exception as e:
        print(f"Unexpected error: {str(e)}")
        return JsonResponse({"error": f"Error processing file: {str(e)}"}, status=500)
    finally:
        # Clean up temporary file
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.unlink(tmp_path)
                print(f"Cleaned up temp file: {tmp_path}")
            except Exception as e:
                print(f"Error deleting temp file: {e}")


SYNCFUSION_SAVE_URLS = [
    # Primary
    "https://document.syncfusion.com/web-services/spreadsheet-editor/api/spreadsheet/save",
    # Fallback
    "https://ej2services.syncfusion.com/production/web-services/api/spreadsheet/save",
]


def save_excel_local(file_path, data):
    wb = load_workbook(file_path)
    ws = wb.active
    ws["A1"] = "Updated locally!"
    wb.save(file_path)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def spreadsheet_save(request):
    """
    Convert spreadsheet JSON back to Excel and replace Cloudinary file.
    """
    document_id = request.data.get("documentId")
    json_data = request.data.get("JSONData")
    file_name = request.data.get("FileName", "Spreadsheet.xlsx")
    save_type = request.data.get("saveType", "Xlsx")

    if not all([document_id, json_data]):
        return JsonResponse({"error": "Missing required data"}, status=400)

    try:
        document = Document.objects.get(id=document_id)
    except Document.DoesNotExist:
        return JsonResponse({"error": "Document not found"}, status=404)

    # Permission check: must be allowed to edit original
    if not document.can_edit(request.user):
        if document.is_in_public_folder():
            return JsonResponse(
                {
                    "error": "Cannot save public documents",
                    "message": "This document is public. Please make a copy to save changes.",
                    "isPublic": True,
                },
                status=403,
            )
        return JsonResponse(
            {
                "error": "No edit permission",
                "message": "Only the owner or admin can edit this document.",
            },
            status=403,
        )

    try:
        # Normalize filename to ensure .xlsx extension
        base_name, ext = os.path.splitext(
            file_name or document.title or "Spreadsheet.xlsx"
        )
        if ext.lower() not in [".xlsx", ".xls"]:
            file_name = f"{base_name}.xlsx"
        else:
            file_name = f"{base_name}{ext}"

        # JSONData may arrive as a string or object; ensure it's a string (per Syncfusion save API)
        try:
            if isinstance(json_data, str):
                # Keep as-is (already a JSON string of the Workbook)
                workbook_json_str = json_data
            else:
                # Convert dict/object to string
                workbook_json_str = json.dumps(json_data)
        except Exception:
            # As a last resort, cast to string
            workbook_json_str = str(json_data)

        last_error = None
        content_bytes = None

        # Try multiple Syncfusion endpoints (primary then fallback)
        for url in SYNCFUSION_SAVE_URLS:
            try:
                # Syncfusion save API expects form body fields (not JSON)
                payload = {
                    "FileName": file_name,
                    "saveType": save_type,
                    "JSONData": workbook_json_str,
                    "PdfLayoutSettings": "{}",
                }
                resp = requests.post(url, data=payload, timeout=30)
                if resp.status_code == 200:
                    content_bytes = resp.content
                    break
                else:
                    last_error = f"{resp.status_code}: {resp.text[:500]}"
            except requests.Timeout:
                last_error = "Syncfusion save timed out"
            except Exception as e:
                last_error = str(e)

        if content_bytes is None:
            return JsonResponse(
                {
                    "error": "Syncfusion conversion failed",
                    "details": last_error,
                },
                status=500,
            )

        # Delete the old file from storage to avoid duplicates
        if document.file:
            try:
                document.file.delete(save=False)
            except Exception as e:
                # Log but continue, we'll overwrite metadata anyway
                print(f"Error deleting old file: {e}")

        # Save new content to Cloudinary via Django storage backend
        document.title = file_name
        document.file.save(file_name, ContentFile(content_bytes), save=False)
        document.save(update_fields=["title", "file"])

        return JsonResponse(
            {
                "success": True,
                "message": f"{file_name} saved successfully",
                "documentId": document.id,
                "fileUrl": document.file.url,
            }
        )

    except Exception as e:
        return JsonResponse({"error": f"Error saving spreadsheet: {e}"}, status=500)
